# Data extraction, transformation and loading (ETL) operations from source files to MongoD collections 
 
import json 
import csv 
import openpyxl as xl 
from pymongo import MongoClient 
from pprint import pprint 
 
mdb_client = MongoClient() 
 
db = mdb_client.uefa.playerTransfers 
db.remove({}) 
 
# Loading player profiles data from excel file 
player_transfer_sheet = xl.load_workbook('data_sources/playerTransfers.xlsx').active 
 
for x in player_transfer_sheet: 
	db.insert_one({"_id": x[8].value, "player": {"ref": "playerprofiles", "id": x[0].value}, "from_club": {"ref": "clubs", "id": x[1].value},
		"to_club": {"ref": "clubs", "id": x[2].value}, "transfer_fee": x[3].value, "from_league": {"ref": "leagues", "id": x[4].value}, 
		 "to_league": {"ref": "leagues", "id": x[5].value}, "agent": {"ref": "agents", "id": x[6].value}, "market_infl": x[7].value}) 
 
print("\n> Inserted Transfer data into UEFA Transfers Collection\n") 
 
db = mdb_client.uefa.leagues 
db.remove({}) 
 
# Loading league profiles data from csv file 
with open('data_sources/LeagueStats.csv') as csv_source: 
	source_data = csv.reader(csv_source) 
	for line in source_data: 
		# print(line)	 
		if line[0] != 'L_Id': 
			db.insert_one({"_id": line[0], "name": line[1], "country": line[1] , "team_count": line[2],  
				"'Avg_Transfer_spending": line[6]}) 
 
print("> Inserted League data into UEFA League Collection\n") 
 
db = mdb_client.uefa.clubs 
db.remove({}) 
 
# Loading club profiles data from excel file 
club_profile_sheet = xl.load_workbook('data_sources/ClubProfile.xlsx').active 
 
for x in club_profile_sheet: 
	# print(x[0].value+str(x[1].value)+x[2].value+x[3].value) 
	if x[3].value != 'Club_Id': 
		db.insert_one({"_id": x[3].value, "name": x[0].value, "budget": x[1].value , "manager": x[2].value, 
	 		"league": {"ref": "leagues", "id": x[4].value}, "global_ranking": x[5].value}) 
 
print("> Inserted Clubs data into UEFA Clubs Collection\n") 
 
db = mdb_client.uefa.agents 
db.remove({}) 
 
# Loading agents profiles data from csv file 
with open('data_sources/AgentProfile.csv') as csv_source: 
	source_data = csv.reader(csv_source) 
	for line in source_data: 
		# print(line)	 
		if line[1] != 'Agent_Id': 
			db.insert_one({"_id": line[1], "name": line[0], "popularity": line[2]}) 
 
print("> Inserted Agents data into UEFA Agent Collection\n") 
 
db = mdb_client.uefa.playerprofiles 
db.remove({}) 
 
 
# Loading player profiles data from excel file 
player_profile_sheet = xl.load_workbook('data_sources/PlayerProfile.xlsx').active 
 
for x in player_profile_sheet: 
	# print(x[0].value+str(x[1].value)+x[2].value+x[3].value) 
	if x[5].value != 'Player_Id': 
		db.insert_one({"_id": x[5].value, "name": x[0].value, "age": x[1].value , "position": x[2].value,  
			"valuation": x[3].value, "rating": x[4].value, "agent": {"ref": "agents", "id": x[6].value}, "club": {"ref": "clubs", "id": x[7].value}}) 
 
print("> Inserted Player data into UEFA Profiles Collection\n") 
 
# for doc in db.find(): 
# 	print(doc) 
 
# print("> Inserted PL data into UEFA Clubs Collection : \n") 
# Loading clubs data from json file 
# with open('uefa_country_clubs.json') as json_source:  
# 	source_data = json.load(json_source) 
# 	# pprint(source_data['tournaments']) 
# 	for x in source_data['tournaments']: 
# 		# pprint(x['name']) 
# 		if x['name'] == "English Premier League 2015/16": 
# 			for y in x['clubs']: 
# 				db.insert_one(y) 